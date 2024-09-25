from openpyxl import load_workbook


workbook = load_workbook(filename="sample.xlsx")
workbook.sheetnames

# file as root node
# sheets as packages

# Sheet column format
# DataSet = Package | Segment = Object | Attribute | Data Type | M / C / O | Definition